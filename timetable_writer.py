import sqlite3
import datetime
import pip
from collections import Counter
import locale
from openpyxl import *


# def install(package):
#     pip.main(['install', package])
#
# try:
#     from openpyxl import *
# except ImportError:
#     install("openpyxl")


class Timetable_writer:

    def __init__(self, template, outfile, from_date, to_date):
        self.from_date = from_date
        self.to_date = to_date
        self.outfile = outfile
        self.path_to_db = "prepod_timetable_base.db"
        self.template = template
        self.connect = sqlite3.connect(self.path_to_db)
        self.cursor = self.connect.cursor()
        locale.setlocale(locale.LC_ALL, "Russian")

    def run(self):
        # получаем список преподавателей

        def select_table_name():
            select_res = self.cursor.execute("SELECT tbl_name FROM sqlite_master").fetchall()
            prepod_list = []
            for item in select_res:
                prepod_list.append(item[0])
            prepod_list.sort()
            return prepod_list

        def format_prepod_name(prepod_name):
            return prepod_name.replace("_", " ")

        def format_date(date_str):
            def swith_month(num_month):
                month = {
                    1: "Января",
                    2: "Февраля",
                    3: "Марта",
                    4: "Апреля",
                    5: "Мая",
                    6: "Июня",
                    7: "Июля",
                    8: "Августа",
                    9: "Сентября",
                    10: "Октября",
                    11: "Ноября",
                    12: "Декабря"
                }
                return month[num_month]

            name_day = datetime.datetime.strptime(date_str, "%d.%m.%Y")
            name_day = name_day.strftime("%A").capitalize()
            date_lst = date_str.split(".")
            out_str = str(int(date_lst[0])) + " " + swith_month(int(date_lst[1])) + "\n" + name_day
            return out_str

        # Открываем книгу-шаблон -> лист 'template'
        wb = load_workbook(self.template)
        work_sheet = wb["template"]

        # Задаём начальную позицию на листе
        position = {"col": 3, "row": 3}
        number_pp = 1

        # Записываем номер п/п, имя препода
        for prepod in select_table_name():
            # Записываем номер п/п
            work_sheet.cell(column=1, row=position["row"], value=number_pp)
            # Записываем имя препода
            work_sheet.cell(column=2, row=position["row"], value=format_prepod_name(prepod))
            prepod_table = self.cursor.execute("SELECT * FROM {} ".format(prepod), ).fetchall()

            date_list = []
            for row in prepod_table:
                if row[0] not in date_list:
                    date_list.append(row[0])

            # Записываем данные
            for one_date in date_list:
                from_date = datetime.datetime.strptime(self.from_date, "%d.%m.%Y")
                to_date = datetime.datetime.strptime(self.to_date, "%d.%m.%Y")
                to_one_date = datetime.datetime.strptime(one_date, "%d.%m.%Y")
                print(from_date)
                print(to_date)
                # print(to_one_date)
                if from_date < to_one_date < to_date:
                    res_one_day = self.cursor.execute("SELECT * FROM {} WHERE date =?".format(prepod),
                                                      [one_date]).fetchall()
                    # print(res_one_day)
                    add_str = ""
                    add_str += format_date(one_date) + ","
                    for one_para in res_one_day:
                        add_str += "\n"
                        add_str += (one_para[1]) + " " + "ауд." + " " + one_para[2]
                    work_sheet.cell(column=position["col"], row=position["row"], value=add_str)
                    position["col"] += 1
                    if position["col"] == 15:
                        position["col"] = 3
                        position["row"] += 1
            if position["col"] != 3:
                position["col"] = 3
                position["row"] += 1
            number_pp += 1

        wb.save(self.outfile)


if __name__ == "__main__":
    reader = Timetable_writer("template.xlsx", "out.xlsx", "01.01.2018", "01.04.2018")
    reader.run()
